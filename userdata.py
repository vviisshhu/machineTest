import openpyxl
from openpyxl import Workbook
import os

FILE_NAME = "users.xlsx"

def create_excel_file():
    if not os.path.exists(FILE_NAME):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name","Email","Phone Number"])
        workbook.save(FILE_NAME)
        
        
def add_user():
    name = input("Enter Name: ")
    email = input("Enter Email: ")
    phone = input("Enter Phone Number: ")
    
    
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active
    sheet.append([name,email,phone])
    workbook.save(FILE_NAME)
    print(f"User '{name}' added successfully!")
    
    
def display_users():
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active
    print("\nstored Users: ")
    for row in sheet.iter_rows(values_only= True):
        print(row)
        

def main():
    create_excel_file()
    while True:
        print("\nMake a choice to proceed:-")
        print("Press '1' to Add users")
        print("Press '2' to Display users")
        print("Press '3' to Exit.")
        choice = input("nPress any key  to  continue: ")
        
        if choice == '1':
            add_user()
        elif choice == '2':
            display_users()
        elif choice == '3':
            print("Exiting..")
        else:
            print("Invalid choice. try again")
            
if __name__ == "__main__" :
    main()
     